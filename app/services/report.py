from __future__ import annotations

import os
import subprocess
import sys
import tempfile
from collections import OrderedDict
from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path

from app.core.database import session_scope
from app.repositories import ActMedServiceRepository, MedServiceRepository, PaymentRepository
from app.services.exceptions import BusinessRuleError


class ReportService:
    def render_services_report(self, date_from: datetime, date_to: datetime) -> Path:
        period_start = self._start_of_day(date_from)
        period_end = self._start_of_day(date_to)
        rows = self._load_services_report_rows(period_start, period_end + timedelta(days=1))
        return self._render_services_report_xlsx(rows, period_start, period_end)

    def render_financial_report(self, date_from: datetime, date_to: datetime) -> Path:
        period_start = self._start_of_day(date_from)
        period_end = self._start_of_day(date_to)
        rows = self._load_financial_report_rows(period_start, period_end + timedelta(days=1))
        return self._render_financial_report_xlsx(rows, period_start, period_end)

    def render_services_matrix_report(self, date_from: datetime, date_to: datetime) -> Path:
        period_start = self._start_of_day(date_from)
        period_end = self._start_of_day(date_to)
        services, rows = self._load_services_matrix_data(period_start, period_end + timedelta(days=1))
        return self._render_services_matrix_xlsx(services, rows, period_start, period_end)

    def open_report(self, path: Path) -> None:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
            return
        if sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
            return
        subprocess.Popen(["xdg-open", str(path)])

    def _load_services_report_rows(self, date_from: datetime, date_to: datetime) -> list[dict]:
        with session_scope() as session:
            rows = ActMedServiceRepository(session).list_service_report_rows(date_from, date_to)
            return [
                {
                    "date": row.act.date,
                    "contract_number": row.act.contract.contract_number if row.act.contract else "",
                    "service": row.current_name,
                    "price": self._discounted_price(row.price, row.discount),
                    "count": Decimal(str(row.count or 0)),
                }
                for row in rows
            ]

    def _load_financial_report_rows(self, date_from: datetime, date_to: datetime) -> list[dict]:
        with session_scope() as session:
            rows = PaymentRepository(session).list_financial_report_rows(date_from, date_to)
            return [
                {
                    "date": row.date,
                    "contract_number": row.contract.contract_number if row.contract else "",
                    "patient": row.contract.patient_name if row.contract else "",
                    "amount": Decimal(str(row.amount or 0)),
                    "category": row.contract.category if row.contract else "",
                    "comments": row.comments or "",
                }
                for row in rows
            ]

    def _load_services_matrix_data(self, date_from: datetime, date_to: datetime) -> tuple[list[dict], list[dict]]:
        with session_scope() as session:
            services = [
                {"id": service.id, "name": service.name}
                for service in MedServiceRepository(session).list(is_folder=False, limit=None)
            ]
            rows = ActMedServiceRepository(session).list_service_report_rows(date_from, date_to)
            report_rows = [
                {
                    "date": row.act.date,
                    "service_id": row.med_service_id,
                    "service": row.current_name,
                    "count": Decimal(str(row.count or 0)),
                    "cost": self._discounted_price(row.price, row.discount) * Decimal(str(row.count or 0)),
                }
                for row in rows
            ]
            return sorted(services, key=lambda item: item["name"]), report_rows

    def _render_services_report_xlsx(self, rows: list[dict], date_from: datetime, date_to: datetime) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise BusinessRuleError("Для отчётов XLSX установите зависимость openpyxl.") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Услуги"

        grouped_rows = self._group_report_rows(rows)
        title = f"Отчет по услугам за период с {date_from.strftime('%d.%m.%Y')} по {self._inclusive_to(date_to).strftime('%d.%m.%Y')}"

        sheet.merge_cells("A1:F1")
        sheet["A1"] = title
        sheet["A1"].font = Font(bold=True, size=12)
        sheet["A1"].alignment = Alignment(horizontal="center")

        headers = ["Дата", "Номер договора", "Услуга", "Цена", "Количество", "Стоимость"]
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=column, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fills = [PatternFill("solid", fgColor="FFF9C4"), PatternFill("solid", fgColor="C8F7F4")]
        total = Decimal("0")
        row_number = 4

        for group_index, (date_value, items) in enumerate(self._iter_report_days(grouped_rows, date_from, date_to)):
            fill = fills[group_index % len(fills)]
            if not items:
                for column, value in enumerate([date_value.strftime("%d.%m.%Y"), "", "", "", "", ""], start=1):
                    cell = sheet.cell(row=row_number, column=column, value=value)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = self._cell_alignment(column)
                row_number += 1
                continue
            for item_index, item in enumerate(items):
                line_total = item["price"] * item["count"]
                total += line_total
                values = [
                    date_value.strftime("%d.%m.%Y") if item_index == 0 else "",
                    item["contract_number"],
                    item["service"],
                    item["price"],
                    int(item["count"]) if item["count"] == int(item["count"]) else item["count"],
                    line_total,
                ]
                for column, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_number, column=column, value=value)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = self._cell_alignment(column)
                    if column in {4, 6}:
                        cell.number_format = '#,##0.00'
                row_number += 1

        total_row = row_number
        sheet.cell(row=total_row, column=5, value="Итого:")
        sheet.cell(row=total_row, column=6, value=total)
        for column in range(5, 7):
            cell = sheet.cell(row=total_row, column=column)
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = self._cell_alignment(column)
            if column == 6:
                cell.number_format = '#,##0.00'

        widths = [14, 18, 70, 16, 12, 16]
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.freeze_panes = "A4"

        output_dir = Path(tempfile.gettempdir()) / "rd4"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"services_report_{date_from.strftime('%Y%m%d')}_{self._inclusive_to(date_to).strftime('%Y%m%d')}.xlsx"
        workbook.save(output_path)
        return output_path

    def _render_services_matrix_xlsx(
        self,
        services: list[dict],
        rows: list[dict],
        date_from: datetime,
        date_to: datetime,
    ) -> Path:
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise BusinessRuleError("Для отчётов XLSX установите зависимость openpyxl.") from exc

        dates = [date_value for date_value, _items in self._iter_report_days(OrderedDict(), date_from, date_to)]
        service_index = self._services_with_report_rows(services, rows)
        metrics = self._build_services_matrix_metrics(rows)

        workbook = Workbook()
        quantity_sheet = workbook.active
        quantity_sheet.title = "Количество"
        cost_sheet = workbook.create_sheet("Стоимость")

        self._fill_services_matrix_sheet(
            quantity_sheet,
            f"Отчет по услугам за период с {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')}: количество",
            service_index,
            dates,
            metrics,
            "count",
            "0",
        )
        self._fill_services_matrix_sheet(
            cost_sheet,
            f"Отчет по услугам за период с {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')}: стоимость",
            service_index,
            dates,
            metrics,
            "cost",
            '#,##0.00',
        )

        output_dir = Path(tempfile.gettempdir()) / "rd4"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"services_matrix_report_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
        workbook.save(output_path)
        return output_path

    def _render_financial_report_xlsx(self, rows: list[dict], date_from: datetime, date_to: datetime) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise BusinessRuleError("Для отчётов XLSX установите зависимость openpyxl.") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Финансы"

        grouped_rows = self._group_financial_report_rows(rows)
        title = f"Финансовый отчет за период с {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')}"

        sheet.merge_cells("A1:F1")
        sheet["A1"] = title
        sheet["A1"].font = Font(bold=True, size=12)
        sheet["A1"].alignment = Alignment(horizontal="center")

        headers = ["Дата", "№ договора", "Пациент", "Сумма", "Категория", "Комментарии"]
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=column, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fills = [PatternFill("solid", fgColor="FFF9C4"), PatternFill("solid", fgColor="D8FAD5")]
        total = Decimal("0")
        category_totals: OrderedDict[str, Decimal] = OrderedDict()
        row_number = 4

        for group_index, (date_value, items) in enumerate(self._iter_report_days(grouped_rows, date_from, date_to)):
            fill = fills[group_index % len(fills)]
            if not items:
                for column, value in enumerate([date_value.strftime("%d.%m.%Y"), "", "", "", "", ""], start=1):
                    cell = sheet.cell(row=row_number, column=column, value=value)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = self._financial_cell_alignment(column)
                row_number += 1
                continue

            for item_index, item in enumerate(items):
                amount = item["amount"]
                total += amount
                category = item["category"] or "Без категории"
                category_totals[category] = category_totals.get(category, Decimal("0")) + amount
                values = [
                    date_value.strftime("%d.%m.%Y") if item_index == 0 else "",
                    item["contract_number"],
                    item["patient"],
                    amount,
                    item["category"],
                    item["comments"],
                ]
                for column, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_number, column=column, value=value)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = self._financial_cell_alignment(column)
                    if column == 4:
                        cell.number_format = '#,##0.00'
                row_number += 1

        total_row = row_number
        sheet.cell(row=total_row, column=5, value="Итого:")
        sheet.cell(row=total_row, column=6, value=total)
        self._style_financial_total_row(sheet, total_row, border)
        row_number += 1

        for category, amount in category_totals.items():
            sheet.cell(row=row_number, column=5, value=f"Итого {category}:")
            sheet.cell(row=row_number, column=6, value=amount)
            self._style_financial_total_row(sheet, row_number, border)
            row_number += 1

        widths = [14, 18, 34, 16, 18, 56]
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.freeze_panes = "A4"

        output_dir = Path(tempfile.gettempdir()) / "rd4"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"financial_report_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
        workbook.save(output_path)
        return output_path

    def _group_report_rows(self, rows: list[dict]) -> OrderedDict:
        grouped: OrderedDict = OrderedDict()
        for row in rows:
            date_key = row["date"].date()
            grouped.setdefault(date_key, OrderedDict())
            service_key = (row["contract_number"], row["service"], row["price"])
            if service_key not in grouped[date_key]:
                grouped[date_key][service_key] = {
                    "contract_number": row["contract_number"],
                    "service": row["service"],
                    "price": row["price"],
                    "count": Decimal("0"),
                }
            grouped[date_key][service_key]["count"] += row["count"]
        return OrderedDict((date_key, list(items.values())) for date_key, items in grouped.items())

    def _group_financial_report_rows(self, rows: list[dict]) -> OrderedDict:
        grouped: OrderedDict = OrderedDict()
        for row in rows:
            date_key = row["date"].date()
            grouped.setdefault(date_key, [])
            grouped[date_key].append(row)
        return grouped

    def _services_with_report_rows(self, services: list[dict], rows: list[dict]) -> list[dict]:
        result: OrderedDict[int, dict] = OrderedDict((service["id"], service) for service in services)
        for row in rows:
            if row["service_id"] not in result:
                result[row["service_id"]] = {"id": row["service_id"], "name": row["service"]}
        return list(result.values())

    def _build_services_matrix_metrics(self, rows: list[dict]) -> dict:
        metrics: dict[tuple[int, object], dict[str, Decimal]] = {}
        for row in rows:
            key = (row["service_id"], row["date"].date())
            if key not in metrics:
                metrics[key] = {"count": Decimal("0"), "cost": Decimal("0")}
            metrics[key]["count"] += row["count"]
            metrics[key]["cost"] += row["cost"]
        return metrics

    def _discounted_price(self, price, discount) -> Decimal:
        price = Decimal(str(price or 0))
        discount = Decimal(str(discount or 0))
        return (price * (Decimal("1") - discount / Decimal("100"))).quantize(Decimal("0.01"))

    def _iter_report_days(self, grouped_rows: OrderedDict, date_from: datetime, date_to: datetime):
        current = date_from.date()
        last = date_to.date()
        while current <= last:
            yield current, grouped_rows.get(current, [])
            current = current + timedelta(days=1)

    def _inclusive_to(self, date_to: datetime) -> datetime:
        return date_to

    def _start_of_day(self, value: datetime) -> datetime:
        return value.replace(hour=0, minute=0, second=0, microsecond=0)

    def _cell_alignment(self, column: int):
        try:
            from openpyxl.styles import Alignment
        except ImportError as exc:
            raise BusinessRuleError("Для отчётов XLSX установите зависимость openpyxl.") from exc
        if column == 3:
            return Alignment(horizontal="left", vertical="center", wrap_text=True)
        if column in {4, 5, 6}:
            return Alignment(horizontal="right", vertical="center")
        return Alignment(horizontal="center", vertical="center")

    def _fill_services_matrix_sheet(
        self,
        sheet,
        title: str,
        services: list[dict],
        dates: list,
        metrics: dict,
        metric_name: str,
        number_format: str,
    ) -> None:
        try:
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise BusinessRuleError("Для отчётов XLSX установите зависимость openpyxl.") from exc

        last_column = len(dates) + 2
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        sheet.cell(row=1, column=1, value=title)
        sheet.cell(row=1, column=1).font = Font(bold=True, size=12)
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="F3F4F6")
        day_fills = [PatternFill("solid", fgColor="FFF9C4"), PatternFill("solid", fgColor="D8FAD5")]

        sheet.cell(row=3, column=1, value="Услуга")
        sheet.cell(row=3, column=last_column, value="Итого")
        for index, date_value in enumerate(dates, start=2):
            cell = sheet.cell(row=3, column=index, value=date_value.strftime("%d.%m.%Y"))
            cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
            cell.fill = day_fills[(index - 2) % len(day_fills)]
        for column in range(1, last_column + 1):
            cell = sheet.cell(row=3, column=column)
            cell.font = Font(bold=True)
            if column in {1, last_column}:
                cell.fill = header_fill
            cell.border = border
            if column in {1, last_column}:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        row_number = 4
        for service in services:
            row_total = Decimal("0")
            service_name_cell = sheet.cell(row=row_number, column=1, value=service["name"])
            service_name_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            service_name_cell.border = border
            for column, date_value in enumerate(dates, start=2):
                value = metrics.get((service["id"], date_value), {}).get(metric_name, Decimal("0"))
                row_total += value
                cell = sheet.cell(row=row_number, column=column, value=self._matrix_value(value))
                cell.fill = day_fills[(column - 2) % len(day_fills)]
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = number_format
            total_cell = sheet.cell(row=row_number, column=last_column, value=self._matrix_value(row_total))
            total_cell.border = border
            total_cell.font = Font(bold=True)
            total_cell.alignment = Alignment(horizontal="center", vertical="center")
            total_cell.number_format = number_format
            row_number += 1

        sheet.column_dimensions["A"].width = 72
        for column in range(2, last_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 12
        sheet.row_dimensions[3].height = 70
        sheet.freeze_panes = "B4"

    def _matrix_value(self, value: Decimal):
        if value == int(value):
            return int(value)
        return value

    def _financial_cell_alignment(self, column: int):
        try:
            from openpyxl.styles import Alignment
        except ImportError as exc:
            raise BusinessRuleError("Для отчётов XLSX установите зависимость openpyxl.") from exc
        if column in {3, 6}:
            return Alignment(horizontal="left", vertical="center", wrap_text=True)
        if column == 4:
            return Alignment(horizontal="right", vertical="center")
        return Alignment(horizontal="center", vertical="center")

    def _style_financial_total_row(self, sheet, row_number: int, border) -> None:
        for column in range(5, 7):
            cell = sheet.cell(row=row_number, column=column)
            cell.font = self._bold_font()
            cell.border = border
            cell.alignment = self._financial_cell_alignment(column)
            if column == 6:
                cell.number_format = '#,##0.00'

    def _bold_font(self):
        try:
            from openpyxl.styles import Font
        except ImportError as exc:
            raise BusinessRuleError("Для отчётов XLSX установите зависимость openpyxl.") from exc
        return Font(bold=True)
