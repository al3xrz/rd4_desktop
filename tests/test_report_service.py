from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import load_workbook

from tests.test_models_smoke import configure_temp_database
from tests.test_services_smoke import contract_payload


def test_services_report_xlsx_groups_rows_and_totals(tmp_path, monkeypatch):
    configure_temp_database(tmp_path, monkeypatch)

    from app.models import Role
    from app.services import ActService, AuthService, ContractService, MedServiceService, ReportService

    now = datetime(2026, 5, 13, 23, 59, tzinfo=timezone.utc)
    admin = AuthService().create_user({"username": "admin", "password": "secret", "role": Role.ADMIN})
    contract = ContractService().create_contract(contract_payload(now), admin)
    folder = MedServiceService().create_folder({"name": "Root"})
    service = MedServiceService().create_service(
        {
            "parent_id": folder.id,
            "code": "A01",
            "name": "Consultation",
            "unit": "шт",
            "price": Decimal("100.00"),
            "vat": 0,
        }
    )
    other_service = MedServiceService().create_service(
        {
            "parent_id": folder.id,
            "code": "A02",
            "name": "Diagnostics",
            "unit": "шт",
            "price": Decimal("50.00"),
            "vat": 0,
        }
    )

    ActService().create_act(
        contract.id,
        {
            "number": "A-001",
            "date": now,
            "services": [
                {"med_service_id": service.id, "count": 1},
                {"med_service_id": service.id, "count": 2},
                {"med_service_id": other_service.id, "count": 1, "discount": Decimal("10.00")},
            ],
        },
        admin,
    )

    date_from = datetime(2026, 5, 12, 12, 30, tzinfo=timezone.utc)
    date_to = datetime(2026, 5, 13, 1, 15, tzinfo=timezone.utc)
    path = ReportService().render_services_report(date_from, date_to)
    workbook = load_workbook(path)
    sheet = workbook.active

    assert sheet["A1"].value == "Отчет по услугам за период с 12.05.2026 по 13.05.2026"
    assert sheet["A3"].value == "Дата"
    assert sheet["B3"].value == "Номер договора"
    assert sheet["C3"].value == "Услуга"
    assert sheet["A4"].value == "12.05.2026"
    assert sheet["B4"].value is None
    assert sheet["A5"].value == "13.05.2026"
    assert sheet["B5"].value == "C-001"
    assert sheet["C5"].value == "Consultation"
    assert sheet["C5"].alignment.horizontal == "left"
    assert sheet["E5"].value == 3
    assert sheet["F5"].value == 300
    assert sheet["C6"].value == "Diagnostics"
    assert sheet["D6"].value == 45
    assert sheet["F6"].value == 45
    assert sheet["E7"].value == "Итого:"
    assert sheet["F7"].value == 345


def test_financial_report_xlsx_includes_empty_days_and_category_totals(tmp_path, monkeypatch):
    configure_temp_database(tmp_path, monkeypatch)

    from app.models import Role
    from app.services import AuthService, ContractService, PaymentService, ReportService

    first_day = datetime(2026, 5, 11, 10, 30, tzinfo=timezone.utc)
    third_day = datetime(2026, 5, 13, 23, 59, tzinfo=timezone.utc)
    admin = AuthService().create_user({"username": "admin", "password": "secret", "role": Role.ADMIN})

    first_payload = contract_payload(first_day)
    first_payload["category"] = "Категория 1"
    first_payload["patient_name"] = "First Patient"
    first_payload["prepay_inpatient_treatment"] = Decimal("0")
    first_contract = ContractService().create_contract(first_payload, admin)

    second_payload = contract_payload(first_day)
    second_payload["contract_number"] = "C-002"
    second_payload["category"] = "Категория 2"
    second_payload["patient_name"] = "Second Patient"
    second_payload["prepay_inpatient_treatment"] = Decimal("0")
    second_contract = ContractService().create_contract(second_payload, admin)

    payments = PaymentService()
    payments.create_payment(first_contract.id, {"date": first_day, "amount": Decimal("100.00"), "comments": "Оплата"}, admin)
    payments.create_payment(second_contract.id, {"date": third_day, "amount": Decimal("50.00"), "comments": "Доплата"}, admin)

    path = ReportService().render_financial_report(
        datetime(2026, 5, 11, 18, 0, tzinfo=timezone.utc),
        datetime(2026, 5, 13, 1, 0, tzinfo=timezone.utc),
    )
    workbook = load_workbook(path)
    sheet = workbook.active

    assert sheet["A1"].value == "Финансовый отчет за период с 11.05.2026 по 13.05.2026"
    assert sheet["B3"].value == "№ договора"
    assert sheet["C3"].value == "Пациент"
    assert sheet["A4"].value == "11.05.2026"
    assert sheet["B4"].value == "C-001"
    assert sheet["C4"].value == "First Patient"
    assert sheet["D4"].value == 100
    assert sheet["E4"].value == "Категория 1"
    assert sheet["F4"].value == "Оплата"
    assert sheet["A5"].value == "12.05.2026"
    assert sheet["B5"].value is None
    assert sheet["A6"].value == "13.05.2026"
    assert sheet["B6"].value == "C-002"
    assert sheet["D6"].value == 50
    assert sheet["A4"].fill.fgColor.rgb != sheet["A5"].fill.fgColor.rgb
    assert sheet["A4"].fill.fgColor.rgb == sheet["B4"].fill.fgColor.rgb
    assert sheet["A6"].fill.fgColor.rgb == sheet["A4"].fill.fgColor.rgb
    assert sheet["E7"].value == "Итого:"
    assert sheet["F7"].value == 150
    assert sheet["E8"].value == "Итого Категория 1:"
    assert sheet["F8"].value == 100
    assert sheet["E9"].value == "Итого Категория 2:"
    assert sheet["F9"].value == 50


def test_reports_ignore_soft_deleted_act_rows_and_payments(tmp_path, monkeypatch):
    configure_temp_database(tmp_path, monkeypatch)

    from app.models import Role
    from app.services import ActService, AuthService, ContractService, MedServiceService, ReportService

    now = datetime(2026, 5, 13, 12, 0, tzinfo=timezone.utc)
    admin = AuthService().create_user({"username": "admin", "password": "secret", "role": Role.ADMIN})
    payload = contract_payload(now)
    payload["prepay_inpatient_treatment"] = Decimal("0")
    contract = ContractService().create_contract(payload, admin)
    folder = MedServiceService().create_folder({"name": "Root"})
    service = MedServiceService().create_service(
        {
            "parent_id": folder.id,
            "code": "A01",
            "name": "Consultation",
            "unit": "шт",
            "price": Decimal("100.00"),
            "vat": 0,
        }
    )

    acts = ActService()
    act = acts.create_act(
        contract.id,
        {
            "number": "A-DELETED",
            "date": now,
            "services": [{"med_service_id": service.id, "count": 2}],
        },
        admin,
        add_payment=True,
    )
    acts.delete_act(act.id, admin)

    reports = ReportService()
    date_from = datetime(2026, 5, 13, tzinfo=timezone.utc)
    date_to = datetime(2026, 5, 13, tzinfo=timezone.utc)

    services_workbook = load_workbook(reports.render_services_report(date_from, date_to))
    services_sheet = services_workbook.active
    assert services_sheet["A4"].value == "13.05.2026"
    assert services_sheet["B4"].value is None
    assert services_sheet["E5"].value == "Итого:"
    assert services_sheet["F5"].value == 0

    financial_workbook = load_workbook(reports.render_financial_report(date_from, date_to))
    financial_sheet = financial_workbook.active
    assert financial_sheet["A4"].value == "13.05.2026"
    assert financial_sheet["B4"].value is None
    assert financial_sheet["E5"].value == "Итого:"
    assert financial_sheet["F5"].value == 0

    matrix_workbook = load_workbook(reports.render_services_matrix_report(date_from, date_to))
    quantity = matrix_workbook["Количество"]
    cost = matrix_workbook["Стоимость"]
    rows_by_name = {quantity.cell(row=row, column=1).value: row for row in range(4, 5)}
    service_row = rows_by_name["Consultation"]
    assert quantity.cell(row=service_row, column=2).value == 0
    assert quantity.cell(row=service_row, column=3).value == 0
    assert cost.cell(row=service_row, column=2).value == 0
    assert cost.cell(row=service_row, column=3).value == 0


def test_services_matrix_report_has_quantity_and_cost_sheets(tmp_path, monkeypatch):
    configure_temp_database(tmp_path, monkeypatch)

    from app.models import Role
    from app.services import ActService, AuthService, ContractService, MedServiceService, ReportService

    now = datetime(2026, 5, 13, 15, 0, tzinfo=timezone.utc)
    admin = AuthService().create_user({"username": "admin", "password": "secret", "role": Role.ADMIN})
    contract = ContractService().create_contract(contract_payload(now), admin)
    med_services = MedServiceService()
    folder = med_services.create_folder({"name": "Root"})
    consultation = med_services.create_service(
        {
            "parent_id": folder.id,
            "code": "A01",
            "name": "Consultation",
            "unit": "шт",
            "price": Decimal("100.00"),
            "vat": 0,
        }
    )
    diagnostics = med_services.create_service(
        {
            "parent_id": folder.id,
            "code": "A02",
            "name": "Diagnostics",
            "unit": "шт",
            "price": Decimal("50.00"),
            "vat": 0,
        }
    )
    med_services.create_service(
        {
            "parent_id": folder.id,
            "code": "A03",
            "name": "Unused",
            "unit": "шт",
            "price": Decimal("10.00"),
            "vat": 0,
        }
    )

    ActService().create_act(
        contract.id,
        {
            "number": "A-MATRIX",
            "date": now,
            "services": [
                {"med_service_id": consultation.id, "count": 3},
                {"med_service_id": diagnostics.id, "count": 2, "discount": Decimal("10.00")},
            ],
        },
        admin,
    )

    path = ReportService().render_services_matrix_report(
        datetime(2026, 5, 12, tzinfo=timezone.utc),
        datetime(2026, 5, 13, tzinfo=timezone.utc),
    )
    workbook = load_workbook(path)
    quantity = workbook["Количество"]
    cost = workbook["Стоимость"]

    assert quantity["A1"].value == "Отчет по услугам за период с 12.05.2026 по 13.05.2026: количество"
    assert quantity["B3"].value == "12.05.2026"
    assert quantity["C3"].value == "13.05.2026"
    assert quantity["D3"].value == "Итого"

    rows_by_name = {quantity.cell(row=row, column=1).value: row for row in range(4, 7)}
    consultation_row = rows_by_name["Consultation"]
    diagnostics_row = rows_by_name["Diagnostics"]
    unused_row = rows_by_name["Unused"]

    assert quantity.cell(row=consultation_row, column=2).value == 0
    assert quantity.cell(row=consultation_row, column=3).value == 3
    assert quantity.cell(row=consultation_row, column=4).value == 3
    assert quantity.cell(row=diagnostics_row, column=3).value == 2
    assert quantity.cell(row=unused_row, column=4).value == 0
    assert quantity["B4"].fill.fgColor.rgb != quantity["C4"].fill.fgColor.rgb
    assert quantity["B4"].fill.fgColor.rgb == quantity["B5"].fill.fgColor.rgb

    assert cost.cell(row=consultation_row, column=3).value == 300
    assert cost.cell(row=consultation_row, column=4).value == 300
    assert cost.cell(row=diagnostics_row, column=3).value == 90
    assert cost.cell(row=unused_row, column=4).value == 0
